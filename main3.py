from collections import Counter

from tkinter import *

import time

import matplotlib.pyplot as plt

from mpl_toolkits.mplot3d import Axes3D

import sys

from math import *

from openpyxl import load_workbook

from inspect import currentframe, getframeinfo

import time

import math

import statistics

import numpy as np

cf = currentframe()

w2 = load_workbook("output.xlsx")

sheet2 = w2.active

w = load_workbook("infos.xlsx")

sheet = w.active

coeff_du_requa = []

formule_chem = sheet.cell(row=1, column=1).value

r_equa = []

r_equa2 = []

t = 0

T = 0

T2 = 0

reactifs = []

reactifs2 = []

r2 = str

reactifs_f = []

passe = 0

produits = []

produits2 = []

p1 = str

produits_f = []

if len(formule_chem) > 1:

    formule_chem = " ".join(formule_chem)

    spll = str.split(formule_chem)

    while t < len(spll):

        if passe == 0:
            reactifs.insert(t, spll[t])

        if passe == 1:
            produits.insert(t, spll[t])

        if spll[t] == "=":
            passe = 1

        t = t + 1

    t = 0

    while t < len(reactifs):

        r2 = "".join(reactifs2)

        if reactifs[t] == "+" or reactifs[t] == "-" or t + 1 == len(reactifs):

            reactifs_f.insert(t, r2)

            reactifs2 = []

        else:

            reactifs2.insert(t, reactifs[t])

        t = t + 1

    t = 0

    while t < len(produits):

        p2 = "".join(produits2)

        if produits[t] == "+" or produits[t] == "-" or t + 1 == len(produits):

            if t + 1 == len(produits):

                produits2.insert(t, produits[t])

            p2 = "".join(produits2)

            produits_f.insert(t, p2)

            produits2 = []

        else:

            produits2.insert(t, produits[t])

        t = t + 1

    t = 0

    while T < len(reactifs_f) + len(produits_f):

        name_list = ["H", "He", "Li", "Be", "B", "C", "N", "O", "F", "Ne", "Na", "Mg", "Al", "Si", "P", "S", "Cl", "Ar",
                     "K"
            , "Ca", "Sc", "Ti", "V", "Cr", "Mn", "Fe", "Co", "Ni", "Cu", "Zn", "Ga", "Ge", "As", "Se", "Br", "Kr", "Rb",
                     "Sr",
                     "Y", "Zr", "Nb", "Mo", "Tc", "Ru", "Rh", "Pd", "Ag", "Cd", "In", "Sn", "Sb", "Te", "I", "Xe", "Cs",
                     "Ba", "Hf"
            , "Ta", "W", "Re", "Os", "Ir", "Pt", "Au", "Hg", "Ti", "Pb", "Bi", "Po", "At", "Rn", "Fr", "Ra", "Rf", "Db",
                     "Sg", "Bh", "Hs", "Mt", "Ds", "Rg", "Cn", "Nh", "Fl", "Mc", "Lv", "Ts", "Og", "La", "Ce", "Pr",
                     "Nd",
                     "Pm",
                     "Sm", "Eu", "Gd", "Tb", "Dy", "Ho", "Er", "Tm", "Yb", "Lu", "Ac", "Th", "Pa", "U", "Np", "Pu",
                     "Am",
                     "Cm",
                     "Bk", "Cf", "Es", "Fm", "Md", "No", "Lr"]

        M_list = [1, 4, 7, 9, 11, 12, 14, 16, 19, 20, 23, 24, 27, 28, 31, 32, 35.5, 40, 39, 40, 45, 48, 51, 52, 55, 56
            , 59, 59, 64, 65, 70, 73, 75, 79, 80, 84, 85, 88, 89, 91, 93, 96, 98, 101, 103, 106, 108, 112, 115, 119,
                  122,
                  128, 127, 131, 133, 137, 178, 181, 184, 186, 190, 192, 195, 197, 201, 204, 207, 209, 209, 210, 222,
                  223,
                  226, 267, 268, 269, 270, 269, 278, 281, 281, 285, 286, 289, 289, 293, 294, 294, 139, 140, 141, 144,
                  145,
                  150, 152, 157, 159, 163, 165, 167, 169, 173, 175, 227, 232, 231, 238, 237, 244, 243, 247, 247, 251,
                  252,
                  257, 258, 259, 262]

        name_numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]

        value_numbers = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]

        crochet_name1 = ["(", "["]

        crochet_name2 = [")", "]"]

        par_cro_mult_list = [0]

        par_cro_numb2 = 0

        par_cro_mult_numb = -1

        para_cro_coeff_list = []

        ti = 0

        ti2 = 0

        t_crochet2 = 0

        t_crochet2bis = 0

        t_crochet2bis2 = 0

        t_crochet2bis2_2 = 0

        t_crochet2_list = []

        fois_list = []

        fois_list2 = []

        t_fois2 = 0

        t_crochet = 0

        para_cro_passage = 0

        M_list_global = []

        t_global = 0

        M_list_global2 = []

        t_list_global = 0

        par_cro_i = 0

        par_cro_i2 = 1

        pp_element2 = 0

        suivant = 0

        zero = 0

        coeff_mult = 0

        coeff_stoechio = 0

        mult_coeff_stoechio = 1

        coeff_stoechio_passage = 0

        coeff_mult_passage = 1

        par_cro_mult_list_second = []

        t_crochet2bis3 = 0

        t_fois = 0

        M_element_chem = 0

        M_element_chem2 = 0

        M_coeff_mult = 0

        if T < len(reactifs_f):

            molecule = reactifs_f[T]

        else:

            molecule = produits_f[T2]

            T2 = T2 + 1

        molecule = " ".join(molecule)

        spl = str.split(molecule)

        spl2 = []

        t_spl = 0

        passagelist1 = []

        passage_numb = 0

        t4_spl = 0

        t3_spl = 0

        t_f = 0

        t5bis = 1

        po = -1

        t_fbis = 0

        t_flist = []

        tpassage = 0

        passage_entre_list = []

        list_eliminate = []

        il_y_a_une_para = 0

        while t_spl < len(spl):

            if spl[t_spl] == "(" or spl[t_spl] == "[":

                list_eliminate.insert(0, 1)

                passage_entre_list.insert(0, 0)

                t5bis = 1

                po = -1

                spl2.insert(len(spl) + 1, t_f)

                passagelist1.insert(t_f, 0)

                t_fbis = 0

                t_flist.insert(t_f, t_f)

                t_f = t_f + 1

                tpassage = tpassage + 1

                while t3_spl < len(passagelist1):

                    passagelist1[t3_spl] = passagelist1[t3_spl] + 1

                    t3_spl = t3_spl + 1

                t3_spl = 0

            else:

                il_y_a_une_para = 1

            if spl[t_spl] == ")" or spl[t_spl] == "]":

                while t3_spl < len(passagelist1):

                    passagelist1[t3_spl] = passagelist1[t3_spl] - 1

                    if passagelist1[t3_spl] == 0:

                        passagelist1.pop()

                        while t4_spl < len(spl2):

                            if spl2[t4_spl] == t_flist[t3_spl]:
                                passage_numb = passage_numb + 1

                            t4_spl = t4_spl + 1

                        if passage_numb > 1:

                            spl2.insert(len(spl) + 1, t_flist[-t5bis])

                            t5bis = t5bis + 1

                            passage_numb = 0

                            t5 = 0

                        else:

                            po = -1

                            t5bis = 1

                            spl2.insert(len(spl) + 1, t_flist[t3_spl])

                            passage_numb = 0

                        t4_spl = 0

                        t_fbis = t_fbis + 1

                    t3_spl = t3_spl + 1

                t3_spl = 0

            else:

                if il_y_a_une_para == 1:

                    il_y_a_une_para = 2

            if il_y_a_une_para == 2:

                spl2.insert(t_spl, spl[t_spl])

            il_y_a_une_para = 0

            t_spl = t_spl + 1

        t_spl = 0

        while t_spl < len(spl2):

            if type(spl2[t_spl]) == int:

                spl2[t_spl] = spl2[t_spl] + 1

            t_spl = t_spl + 1

        l = spl2

        ti = 0

        ti2 = 0

        l22 = []

        l33 = []

        l33b = []

        l4 = []

        l3 = []

        l5 = []

        while ti < len(l):

            if type(l[ti]) != str:

                l22.insert(ti2, l[ti])

                ti2 = ti2 + 1

            ti = ti + 1

        ti = 0

        while ti < len(l22) / 2 + 1:

            l33.insert(ti, ti)

            ti = ti + 1

        l33.remove(l33[0])

        l4 = l33

        l2 = []

        t = 0

        t2 = 0

        t3 = 0

        t4 = 0

        while t < len(l4):

            l3.insert(0, 1)

            l5.insert(0, 0)

            t = t + 1

        t = 0

        t5 = 0

        while t < len(l):

            while t2 < t3:

                l2.insert(t, l[t])

                if l2[t2] == l[t2] and type(l2[t2]) != str:

                    l3[l[t2] - 1] = 0

                t2 = t2 + 1

            t2 = 0

            l2 = []

            while t5 < t4:

                if l3[t5] != 0:

                    l5[t5] = l5[t5] + 1

                t5 = t5 + 1

            t5 = 0

            if t4 < len(l4):

                if l[t] == l4[t4]:

                    t4 = t4 + 1

            t = t + 1

            t3 = t3 + 1

        spl_final = []

        M_passé = 0

        t2 = 0

        l = []

        t_entre = 0

        list_entre = []

        pp_element = 0

        trouvé = 0

        min_list = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q", "r",
                    "s", "t", "u", "v", "w", "x", "y", "z"]
        spl2 = []

        t_spl = 0

        t2_spl = 0

        passagelist1 = []

        passage_numb = 0

        t4_spl = 0

        t5_spl = 0

        t3_spl = 0

        t_f = 0

        t5bis = 1

        po = -1

        t_fbis = 0

        t_flist = []

        tpassage = 0

        while t_spl < len(spl):

            if spl[t_spl] == "(" or spl[t_spl] == "[":

                list_entre.insert(0, 0)

                t5bis = 1

                po = -1

                spl2.insert(len(spl) + 1, t_f)

                passagelist1.insert(t_f, 0)

                t_fbis = 0

                t_flist.insert(t_f, t_f)

                t_f = t_f + 1

                tpassage = tpassage + 1

                while t3_spl < len(passagelist1):
                    passagelist1[t3_spl] = passagelist1[t3_spl] + 1

                    t3_spl = t3_spl + 1

                t3_spl = 0

            if spl[t_spl] == ")" or spl[t_spl] == "]":

                while t3_spl < len(passagelist1):

                    passagelist1[t3_spl] = passagelist1[t3_spl] - 1

                    if passagelist1[t3_spl] == 0:

                        passagelist1.pop()

                        while t4_spl < len(spl2):

                            if spl2[t4_spl] == t_flist[t3_spl]:
                                passage_numb = passage_numb + 1

                            t4_spl = t4_spl + 1

                        if passage_numb > 1:

                            spl2.insert(len(spl) + 1, t_flist[-t5bis])

                            t5bis = t5bis + 1

                            passage_numb = 0

                            t5 = 0

                        else:

                            po = -1

                            t5bis = 1

                            spl2.insert(len(spl) + 1, t_flist[t3_spl])

                            passage_numb = 0

                        t4_spl = 0

                        t_fbis = t_fbis + 1

                    t3_spl = t3_spl + 1

                t3_spl = 0

            t_spl = t_spl + 1

        coco_list = []

        element_de_passage = 0

        t2bis = -1

        while t2 < len(spl):

            while spl[t2] != min_list[t2bis] and t2bis < len(min_list) - 1:

                t2bis = t2bis + 1

                if spl[t2] == min_list[t2bis]:
                    spl_final.pop()

                    l.insert(0, spl[t2 - 1])

                    l.insert(1, min_list[t2bis])

                    l = "".join(l)

                    spl_final.insert(t2, l)

                    l = []

                    trouvé = 1

            if trouvé == 0:
                spl_final.insert(t2, spl[t2])

            t2 = t2 + 1

            t2bis = -1

            trouvé = 0

        t2 = 0

        spl = []

        while t2 < len(spl_final):
            spl.insert(t2, spl_final[t2])

            t2 = t2 + 1

        t = 0

        t_coeff = 0

        t_element_chem = 0

        dernier_M = 0

        M_element_chem_coeff = 0

        t_coco = 0

        while t < len(spl):

            while t_element_chem < len(name_list) and spl[t] != name_list[t_element_chem]:
                t_element_chem = t_element_chem + 1

            if t_element_chem == len(name_list):
                t_element_chem = t_element_chem - 1

            if spl[t] == name_list[t_element_chem]:

                pp_element = 1

                element_de_passage = 1

                para_cro_passage = 0

                coeff_stoechio_passage = 0

                if par_cro_mult_numb > -1:

                    coco_list.insert(t_coco, "d")

                    t_coco = t_coco + 1

                    fois_list2.insert(t_fois2, coeff_mult)

                    t_fois2 = t_fois2 + 1

                    t_crochet2bis2 = t_crochet2bis2 + 1

                    t_crochet2bis3 = t_crochet2bis2

                    if par_cro_numb2 == 0:
                        par_cro_mult_list.insert(t_crochet2bis2, -1)

                        t_crochet2bis2 = t_crochet2bis2 + 1

                        par_cro_mult_list_second.insert(t_crochet2bis3, "fini")

                        t_crochet2bis3 = t_crochet2bis3 + 1

                    par_cro_numb2 = par_cro_numb2 + M_list[t_element_chem]

                    par_cro_mult_list.insert(t_crochet2bis2, par_cro_numb2)

                    par_cro_mult_list_second.insert(t_crochet2bis2, M_list[t_element_chem])

                    M_element_chem2 = M_element_chem2 + M_list[t_element_chem]

                    M_list_global2.insert(t_list_global, M_list[t_element_chem])

                    t_list_global = t_list_global + 1

                    M_list_global.insert(t_list_global, "/")

                    t_list_global = t_list_global + 1

                else:

                    M_element_chem = M_element_chem + M_list[t_element_chem]

                dernier_M = t_element_chem

            t_element_chem = 0

            while t_coeff < len(name_numbers) and spl[t] != name_numbers[t_coeff]:
                t_coeff = t_coeff + 1

            if t_coeff == len(name_numbers):
                t_coeff = t_coeff - 1

                coeff_mult = 0

                coeff_mult_passage = 1

                M_coeff_mult = 0

            if spl[t] == name_numbers[t_coeff]:

                if par_cro_mult_numb > -1 or para_cro_passage == 1:

                    if para_cro_passage == 1:

                        par_cro_i = value_numbers[t_coeff]

                        para_cro_coeff_list.insert(ti, par_cro_i)

                        ti = ti + 1

                    else:

                        coeff_mult = coeff_mult * coeff_mult_passage + value_numbers[t_coeff]

                        if pp_element == 1:
                            coeff_mult_passage = 10

                            pp_element = 0

                            pp_element2 = 1

                        M_passé = M_coeff_mult

                        M_coeff_mult = coeff_mult * M_list[dernier_M]

                        if pp_element2 == 1:

                            M_element_chem2 = M_element_chem2 + M_coeff_mult - M_list[dernier_M]

                            pp_element2 = 0

                        else:

                            M_element_chem2 = M_element_chem2 + M_coeff_mult - M_passé

                        M_list_global.insert(t_list_global, name_list[dernier_M])

                        M_list_global2.insert(t_list_global, M_coeff_mult)

                        t_list_global = t_list_global + 1

                else:

                    if t == 0 or coeff_stoechio_passage == 1:

                        coeff_stoechio = coeff_stoechio * mult_coeff_stoechio + value_numbers[t_coeff]

                        if t == 0:
                            mult_coeff_stoechio = mult_coeff_stoechio * 10

                        coeff_stoechio_passage = 1

                    else:

                        coeff_mult = coeff_mult * coeff_mult_passage + value_numbers[t_coeff]

                        if pp_element == 1:
                            coeff_mult_passage = 10

                            pp_element = 0

                            pp_element2 = 1

                        M_passé = M_coeff_mult

                        M_coeff_mult = coeff_mult * M_list[dernier_M]

                        if pp_element2 == 1:

                            M_element_chem = M_element_chem + M_coeff_mult - M_list[dernier_M]

                            pp_element2 = 0

                        else:

                            M_element_chem = M_element_chem + M_coeff_mult - M_passé

            while t_crochet < len(crochet_name1) and spl[t] != crochet_name1[t_crochet]:
                t_crochet = t_crochet + 1

            if t_crochet == len(crochet_name1):
                t_crochet = 0

            if spl[t] == crochet_name1[t_crochet]:
                element_de_passage = 0

                para_cro_passage = 0

                coeff_stoechio_passage = 0

                par_cro_mult_numb = par_cro_mult_numb + 1

                par_cro_numb2 = 0

            t_crochet = 0

            while t_crochet < len(crochet_name2) and spl[t] != crochet_name2[t_crochet]:
                t_crochet = t_crochet + 1

            if t_crochet == len(crochet_name2):
                t_crochet = 0

            if spl[t] == crochet_name2[t_crochet]:
                element_de_passage = 0

                para_cro_coeff_list.insert(ti, -2)

                ti = ti + 1

                para_cro_passage = 1

                par_cro_mult_numb = par_cro_mult_numb - 1

            t_coeff = 0

            t = t + 1

            t_crochet = 0

        ti = 0

        ti2 = 0

        para_cro_coeff_list2 = []

        while ti < len(para_cro_coeff_list):

            if para_cro_coeff_list[ti] != -2:
                para_cro_coeff_list2.insert(ti, ti)

                ti2 = 1

            ti = ti + 1

        ti = 1

        t = 0

        t_par = 0

        if t_crochet2bis2 > 0:

            while t < 100:

                if t < len(par_cro_mult_list_second):

                    if par_cro_mult_list_second[t] == "fini":
                        t_par = 1

                if t + 1 < len(coco_list):

                    if coco_list[t + 1] != "d" and coco_list[t] == "d":
                        M_element_chem = M_element_chem + par_cro_mult_list_second[t + t_par] * coco_list[t + 1]

                        t_par = 0

                        t = t + 1

                        t_coco = t_coco + 1

                t = t + 1

        t = 0

        para_cro_coeff_list99 = []

        while t < len(para_cro_coeff_list):
            para_cro_coeff_list99.insert(t, para_cro_coeff_list[t])

            t = t + 1

        idx = 0

        idx2 = 0

        para_cro_coeff_list88 = []

        para_cro_coeff_list88_f = []

        lesstw = []

        valid = len(para_cro_coeff_list99)

        t = 0

        ttt = 0

        if len(para_cro_coeff_list) > 1:

            while t < valid:

                if t + 1 < len(para_cro_coeff_list99):

                    if para_cro_coeff_list99[t] == -2 and para_cro_coeff_list99[t + 1] == -2:
                        para_cro_coeff_list.insert(ttt + 1, 1)

                        ttt = ttt + 1

                if t + 1 < len(para_cro_coeff_list99):

                    if para_cro_coeff_list99[t + 1] == -2 and t + 2 == len(para_cro_coeff_list99):
                        para_cro_coeff_list.insert(ttt + 2, 1)

                ttt = ttt + 1

                t = t + 1

            t = 0

            while t < len(para_cro_coeff_list):

                if para_cro_coeff_list[t - 1] == -2:
                    para_cro_coeff_list88.insert(idx, para_cro_coeff_list[t])

                    idx = idx + 1

                para_cro_coeff_list[t] = str(para_cro_coeff_list[t])

                para_cro_coeff_list[t - 1] = str(para_cro_coeff_list[t - 1])

                if len(para_cro_coeff_list[t - 1]) == len(para_cro_coeff_list[t]):
                    para_cro_coeff_list[t] = int(para_cro_coeff_list[t])

                    para_cro_coeff_list88.insert(idx, para_cro_coeff_list[t])

                    idx = idx + 1

                para_cro_coeff_list[t] = int(para_cro_coeff_list[t])

                para_cro_coeff_list[t] = int(para_cro_coeff_list[t])

                if para_cro_coeff_list[t] == -2 and t > 0:

                    idx = 0

                    while idx < len(para_cro_coeff_list88):
                        para_cro_coeff_list88[idx] = str(para_cro_coeff_list88[idx])

                        idx = idx + 1

                    idx = 0

                    joined = "".join(para_cro_coeff_list88)

                    joined = int(joined)

                    para_cro_coeff_list88_f.insert(idx2, joined)

                    idx2 = idx2 + 1

                    para_cro_coeff_list88 = []

                if para_cro_coeff_list[t] == -2:
                    lesstw.insert(0, t)

                t = t + 1

            t = lesstw[0] + 1

            para_cro_coeff_list88 = []

            while t < len(para_cro_coeff_list):
                para_cro_coeff_list[t] = str(para_cro_coeff_list[t])

                para_cro_coeff_list88.insert(t, para_cro_coeff_list[t])

                t = t + 1

            joined = "".join(para_cro_coeff_list88)

            joined = int(joined)

            para_cro_coeff_list88_f.insert(idx2, joined)

        else:

            para_cro_coeff_list88_f = []

            para_cro_coeff_list88_f.insert(0, 1)

        idx = 0

        t = 0

        while t < len(M_list_global):

            if M_list_global[t - 1] == M_list_global[t] and M_list_global[t] != "/":
                M_list_global2[t - 1] = 0

            if t < len(M_list_global) - 1:

                if M_list_global[t] == "/" and M_list_global[t + 1] != "/":
                    M_list_global2[t] = 0

            t = t + 1

        t = 0

        lenghtM = len(M_list_global2)

        while t < lenghtM:

            if 0 in M_list_global2:
                idx = M_list_global2.index(0)

                M_list_global2.remove(M_list_global2[idx])

            t = t + 1

        t = 1

        list_f = []

        add_para = 0

        t2 = 0

        t3 = 0

        par_cro_mult_list_second.insert(len(par_cro_mult_list_second) + 1, "fini")

        while t < len(par_cro_mult_list_second):

            if par_cro_mult_list_second[t] != "fini":

                add_para = add_para + M_list_global2[t3]

                t3 = t3 + 1

                if par_cro_mult_list_second[t + 1] == "fini":
                    list_f.insert(t2, add_para)

            else:

                t2 = t2 + 1

                add_para = 0

            t = t + 1

        ajout_list = []

        list_f3 = []

        pre_final_l = []

        t = 0

        spl2_aa = []

        ajout = 0

        particulier = 0

        vrai_coeff = []

        vrai_coeff2 = []

        spl2_aa2 = []

        while t < len(list_f):
            list_f3.insert(0, "wrong")

            pre_final_l.insert(0, 0)

            t = t + 1

        t = 0

        while t < len(spl2):

            spl2_aa.insert(t, spl2[t])

            counter78 = Counter(spl2_aa)

            if t < len(para_cro_coeff_list88_f):
                vrai_coeff.insert(t, para_cro_coeff_list88_f[t])

            if counter78[spl2_aa[t]] > 1:
                spl2_aa2.insert(t, spl2[t])

                vrai_coeff2.insert(t, spl2_aa[t])

            t = t + 1

        t = 0

        t78 = 0

        passed = 0

        saved_t78 = 0

        while t < len(vrai_coeff):

            while t78 < len(vrai_coeff):

                if t + t78 < len(vrai_coeff) and spl2_aa2[t] > spl2_aa2[t + t78] and t78 > 0:
                    vrai_coeff[t] = vrai_coeff[t] * vrai_coeff[t + t78]

                t78 = t78 + 1

            t78 = 0

            passed = 0

            t = t + 1

        t = 0

        while t < len(list_f):
            pre_final_l[t] = list_f[vrai_coeff2[t]] * vrai_coeff[t]

            t = t + 1

        t = 0

        the_final_result = 0

        while t < len(pre_final_l):
            the_final_result = the_final_result + pre_final_l[t]

            t = t + 1

        globalised_M = M_element_chem + the_final_result

        if coeff_stoechio != 0:
            globalised_M = globalised_M * coeff_stoechio

        if T < len(reactifs_f):

            r_equa.insert(T, globalised_M)

        else:

            r_equa2.insert(T, globalised_M)

        coeff_du_requa.insert(T, coeff_stoechio)

        T = T + 1

T2 = 0

accuracy = sheet.cell(row = 2, column=4).value #precision

M = sheet.cell(row=2, column=3).value

M_l = []

M2 = 0

s2 = []

if len(M) > 0:

    M = " ".join(M)

    s = M.split()

    while T2 < len(s):

        s2.insert(T2, s[T2])

        M2 = "".join(s2)

        if s[T2] == "," or T2 + 1 == len(s):

            if s[T2] == ",":

                s2.pop()

                M2 = "".join(s2)

            M_l.insert(T2, M2)

            s2 = []

        T2 = T2 + 1

    T2 = 0

T2 = 0

while T2 < len(M_l):

    if M_l[T2] == "/":

        M_l[T2] = -1

    T2 += 1

T2 = 0

while T2 < len(M_l):

    M_l[T2] = float(M_l[T2])

    T2 = T2 + 1

print("M", M_l, len(M_l))

M_inter_value = sheet.cell(row=2, column=2).value

inter_Ml = []

M_lf = []

T2 = 0

ro = sheet.cell(row=3, column=3).value

ro_l = []

Cm2 = 0

if len(ro) > 0:

    ro = " ".join(ro)

    s = ro.split()

    while T2 < len(s):

        s2.insert(T2, s[T2])

        Cm2 = "".join(s2)

        if s[T2] == "," or T2 + 1 == len(s):

            if s[T2] == ",":
                s2.pop()

                Cm2 = "".join(s2)

            ro_l.insert(T2, Cm2)

            s2 = []

        T2 = T2 + 1

    T2 = 0

T2 = 0

while T2 < len(ro_l):

    if ro_l[T2] == "/":

        ro_l[T2] = -1

    T2 += 1

T2 = 0

while T2 < len(ro_l):
    ro_l[T2] = float(ro_l[T2])

    T2 = T2 + 1

T2 = 0

print("ro", ro_l)

ro_inter_value = sheet.cell(row=3, column=2).value

ro_lf = []

inter_rol = []

n = sheet.cell(row=4, column=3).value

n2 = 0

n_l = []

if len(n) > 0:

    n = " ".join(n)

    s = n.split()

    while T2 < len(s):

        s2.insert(T2, s[T2])

        n2 = "".join(s2)

        if s[T2] == "," or T2 + 1 == len(s):

            if s[T2] == ",":
                s2.pop()

                n2 = "".join(s2)

            n_l.insert(T2, n2)

            s2 = []

        T2 = T2 + 1

    T2 = 0

T2 = 0

while T2 < len(n_l):

    if n_l[T2] == "/":

        n_l[T2] = -1

    T2 += 1

T2 = 0

while T2 < len(n_l):

    n_l[T2] = float(n_l[T2])

    T2 = T2 + 1

T2 = 0

print("n", n_l)

n_lf = []

n_inter_value = sheet.cell(row=4, column=2).value

inter_nl = []

m = sheet.cell(row=5, column=3).value

m_l = []

m2 = 0

if len(m) > 0:

    m = " ".join(m)

    s = m.split()

    while T2 < len(s):

        s2.insert(T2, s[T2])

        m2 = "".join(s2)

        if s[T2] == "," or T2 + 1 == len(s):

            if s[T2] == ",":
                s2.pop()

                m2 = "".join(s2)

            m_l.insert(T2, m2)

            s2 = []

        T2 = T2 + 1

    T2 = 0

T2 = 0

while T2 < len(m_l):

    if m_l[T2] == "/":

        m_l[T2] = -1

    T2 += 1

T2 = 0

while T2 < len(m_l):

    m_l[T2] = float(m_l[T2])

    T2 = T2 + 1

T2 = 0

m_lf = []

print("m", m_l)

m_inter_value = sheet.cell(row=5, column=2).value

inter_ml = []

v = sheet.cell(row=6, column=3).value

v_l = []

v2 = 0

if len(v) > 0:

    v = " ".join(v)

    s = v.split()

    while T2 < len(s):

        s2.insert(T2, s[T2])

        v2 = "".join(s2)

        if s[T2] == "," or T2 + 1 == len(s):

            if s[T2] == ",":
                s2.pop()

                v2 = "".join(s2)

            v_l.insert(T2, v2)

            s2 = []

        T2 = T2 + 1

    T2 = 0

T2 = 0

while T2 < len(v_l):

    if v_l[T2] == "/":

        v_l[T2] = -1

    T2 += 1

T2 = 0

while T2 < len(v_l):

    v_l[T2] = float(v_l[T2])

    T2 = T2 + 1

T2 = 0

v_lf = []

print("v", v_l)

v_inter_value = sheet.cell(row=6, column=2).value

inter_vl = []

Qr = sheet.cell(row=8, column=3).value

Qr_l = []

Qr2 = 0

if len(Qr) > 0:

    Qr = " ".join(Qr)

    s = Qr.split()

    while T2 < len(s):

        s2.insert(T2, s[T2])

        Qr2 = "".join(s2)

        if s[T2] == "," or T2 + 1 == len(s):

            if s[T2] == ",":

                s2.pop()

                Qr2 = "".join(s2)

            Qr_l.insert(T2, Qr2)

            s2 = []

        T2 = T2 + 1

    T2 = 0

T2 = 0

while T2 < len(Qr_l):

    if Qr_l[T2] == "/":

        Qr_l[T2] = -1

    T2 += 1

T2 = 0

while T2 < len(Qr_l):

    Qr_l[T2] = float(Qr_l[T2])

    T2 = T2 + 1

T2 = 0

Qr_lf = []

Qr_inter_value = sheet.cell(row=8, column=2).value

inter_Qrl = []

inter_Qrn = ""

inter_Qrn2 = 0

while T2 < len(Qr_inter_value): 

    if Qr_inter_value[T2] == ".":

        inter_Qrn = inter_Qrn + "." 

        inter_Qrn = "".join(inter_Qrn)

    if name_numbers.count(Qr_inter_value[T2]) > 0:

        inter_Qrn = inter_Qrn + Qr_inter_value[T2]

        inter_Qrn = "".join(inter_Qrn)

    if name_numbers.count(Qr_inter_value[T2 - 1]) > 0 and Qr_inter_value[T2] == "," or name_numbers.count(Qr_inter_value[T2 - 1]) > 0 and Qr_inter_value[T2] == "]":

        inter_Qrl.append(inter_Qrn)

        inter_Qrn = ""

        inter_Qrn2 += 1

    T2 += 1

T2 = 0

while T2 < len(inter_Qrl):

    inter_Qrl[T2] = float(inter_Qrl[T2])

    T2 += 1

T2 = 0

print("Qr", Qr_l)

inter_Qrl = []

C = sheet.cell(row=7, column=3).value

C2 = 0

C_l = []

if len(C) > 0:

    C = " ".join(C)

    s = C.split()

    while T2 < len(s):

        s2.insert(T2, s[T2])

        C2 = "".join(s2)

        if s[T2] == "," or T2 + 1 == len(s):

            if s[T2] == ",":
                s2.pop()

                C2 = "".join(s2)

            C_l.insert(T2, C2)

            s2 = []

        T2 = T2 + 1

    T2 = 0

T2 = 0

time.sleep(1.5)

while T2 < len(C_l):

    if C_l[T2] == "/":

        C_l[T2] = -1

    T2 += 1

T2 = 0

while T2 < len(C_l):

    C_l[T2] = float(C_l[T2])

    T2 = T2 + 1

T2 = 3

C_lf = []

print("C", C_l)

C_inter_value = sheet.cell(row=7, column=2).value

inter_Cl = []

obj_l = [M_l, ro_l, n_l, m_l, v_l, C_l]

obj_l2 = [M_lf, ro_lf, n_lf, m_lf, v_lf, C_lf]

obj_x = sheet.cell(row=3, column=5).value #graphique info

obj_y = []

while sheet.cell(row=T2, column=6).value != None:

    obj_y.append(sheet.cell(row=T2, column=6).value)

    T2 += 1

T2 = 0

print("graph data", obj_x, obj_y)

molecule_y_spl22 = []

T2 = 0

t = 0

t2 = 0

t375 = 0

les_inter = [M_inter_value, ro_inter_value, n_inter_value, m_inter_value, v_inter_value, C_inter_value, Qr_inter_value]

list_des_inter = [inter_Ml, inter_rol, inter_nl, inter_ml, inter_vl, inter_Cl, inter_Qrl]

M_lf2 = []

ro_lf2 = []

n_lf2 = []

m_lf2 = []

v_lf2 = []

C_lf2 = []

while t375 < len(M_l):

    M_lf2.insert(0, -2)

    ro_lf2.insert(0, -2)

    n_lf2.insert(0, -2)

    m_lf2.insert(0, -2)

    v_lf2.insert(0, -2)

    C_lf2.insert(0, -2)

    t375 = t375 + 1

t375 = 0

t_inter = 0

def interfirst():

    list_val_inter = []

    g = les_inter[t_inter]

    t45 = 0

    g = " ".join(g)

    split2 = g.split()

    split2bis = []

    zf = []

    ajout45 = 0

    while t45 < len(split2):

        split2bis.insert(t45, split2[t45])

        if split2bis[-1] == "," or t45 + 1 == len(split2) or split2bis[-1] == "/":

            if t45 < len(split2):

                if split2[t45] != "/":

                    if t45 + 1 != len(split2):

                        split2bis.pop()

                    ajout45 = "".join(split2bis)

                    split2bis = []

                    zf.insert(t45, ajout45)

                else:

                    zf.insert(t45, "/")

                    split2bis = []

        t45 = t45 + 1

    t45 = 0

    while t45 < len(zf):

        if zf[t45] == "":

            zf.remove(zf[t45])

        t45 = t45 + 1

    t45 = 0

    t_list_val_inter = 0

    t45 = 0

    tt = 0

    tt2 = 0

    zf2 = []

    z_tempo = []

    ajout_z = 0

    zff = []

    while tt < len(zf):

        z = " ".join(zf[tt])

        split = str.split(z)

        while tt2 < len(split):

            zf2.insert(len(zf2) + 1, split[tt2])

            tt2 = tt2 + 1

        tt2 = 0

        tt = tt + 1

    tt2 = 0

    tt = 0

    while tt < len(zf2):

        if zf2[tt] != ";" and zf2[tt] != "]" and zf2[tt] != "[":

            z_tempo.insert(tt, zf2[tt])

        else:

            if zf2[tt] != "/":

                ajout_z = "".join(z_tempo)

                zff.insert(tt, ajout_z)

                z_tempo = []

            else:

                zff.insert(tt, -1)

        tt = tt + 1

    tt = 0

    while tt < len(zff):

        if zff[tt] == "/":

            zff[tt] = -1

        if zff[tt] == "":

            zff.remove(zff[tt])

        tt = tt + 1

    tt = 0

    while tt < len(zff):

        if zff[tt] != -1:

            zff[tt] = float(zff[tt])

        list_des_inter[t_inter].insert(tt, zff[tt])

        tt = tt + 1

    tt = 0

while t_inter < len(list_des_inter):

    interfirst()

    t_inter = t_inter + 1

t_inter = 0

molecule_x_bis = 0

molecule_y_bis = 0

list_molecule_unic = []

T3 = 0

T3b = 0

r_equa_bis = []

r_equa_bis_val = []

molecule_x_finale_theo = []

while T3 < len(r_equa) + len(r_equa2):

    if T3 < len(r_equa):

        r_equa_bis.insert(0, r_equa[T3])

        r_equa_bis[0] = str(r_equa_bis[0])

        r_equa_bis_val.insert(T3, r_equa[T3])

    else:

        r_equa_bis.insert(0, r_equa2[T3b])

        r_equa_bis[0] = str(r_equa_bis[0])

        r_equa_bis_val.insert(T3, r_equa2[T3b])

        T3b = T3b + 1

    T3 = T3 + 1

T3 = 0

molecule_x_bis2 = 0

assez = []

t_qr = 0

print("inter_nl", inter_nl)

def n_fun():  

    t2 = len(n_lf) + 1 

    print(molecule_x_bis)

    if Qr_l[molecule_x_bis] != -2 and len(assez) == 0:

        t_qr = 0

        denu = 0

        numu = 0

    if n_l[molecule_x_bis] != -2 and len(assez) == 0:

        n_lf.insert(t2, n_l[molecule_x_bis])

        assez.insert(0, 1)

    if m_l[molecule_x_bis] != -2 and len(assez) == 0:

        if M_l[molecule_x_bis] != -2:

            assez.insert(0, 1)

            n_lf.insert(t2, m_l[molecule_x_bis] / M_l[molecule_x_bis])

        else:

            if len(r_equa_bis) > 0 and len(assez) == 0:

                assez.insert(0, 1)

                n_lf.insert(t2, m_l[molecule_x_bis] / r_equa_bis_val[molecule_x_bis2])

            if ro_l[molecule_x_bis] != -2 and C_l[molecule_x_bis] != -2 and len(assez) == 0:

                assez.insert(0, 1)

                n_lf.insert(t2, C_l[molecule_x_bis] * (m_l[molecule_x_bis] / ro_l[molecule_x_bis]))

    if M_l[molecule_x_bis] != -2 and len(assez) == 0:

        if ro_l[molecule_x_bis] != -2 and v_l[molecule_x_bis] != -2:

            n_lf.insert(t2, (ro_l[molecule_x_bis] * v_l[molecule_x_bis]) / M_l[molecule_x_bis])

            assez.insert(0, 1)

    if C_l[molecule_x_bis] != -2 and v_l[molecule_x_bis] != -2 and len(assez) == 0:

        assez.insert(0, 1)

        n_lf.insert(t2, C_l[molecule_x_bis] * v_l[molecule_x_bis])

def ro_fun():

    t2 = len(ro_lf) + 1

    if ro_l[molecule_x_bis] != -2 and len(assez) == 0:

        ro_lf.insert(t2, ro_l[molecule_x_bis])

        assez.insert(0, 1)

    if m_l[molecule_x_bis] != -2 and v_l[molecule_x_bis] != -2 and len(assez) == 0:

        ro_lf.insert(t2, m_l[molecule_x_bis] / v_l[molecule_x_bis])

        assez.insert(0, 1)

    if n_l[molecule_x_bis] != -2 and M_l[molecule_x_bis] != -2 and v_l[molecule_x_bis] != -2 and len(assez) == 0:

        ro_lf.insert(t2, (n_l[molecule_x_bis] * M_l[molecule_x_bis]) / v_l[molecule_x_bis])

        assez.insert(0, 1)

    if n_l[molecule_x_bis] != -2 and len(r_equa_bis) > 0 and v_l[molecule_x_bis] != -2 and len(assez) == 0:

        ro_lf.insert(t2, (n_l[molecule_x_bis] * r_equa_bis_val[molecule_x_bis2]) / v_l[molecule_x_bis])

        assez.insert(0, 1)

    if n_l[molecule_x_bis] != -2 and M_l[molecule_x_bis] != -2 and C_l[molecule_x_bis] != -2 and len(assez) == 0:

        ro_lf.insert(t2, C_l[molecule_x_bis] * M_l[molecule_x_bis])

        assez.insert(0, 1)

def M_fun():

    t2 = len(M_lf) + 1

    if M_l[molecule_x_bis] != -2 and len(assez) == 0:

        M_lf.insert(t2, M_l[molecule_x_bis])

        assez.insert(0, 1)

    if len(r_equa_bis) > 0 and len(assez) == 0:

        M_lf.insert(t2, r_equa_bis_val[molecule_x_bis2])

        assez.insert(0, 1)

    if n_l[molecule_x_bis] != -2 and m_l[molecule_x_bis] != -2 and len(assez) == 0:

        M_lf.insert(t2, m_l[molecule_x_bis] / n_l[molecule_x_bis])

        assez.insert(0, 1)

    if ro_l[molecule_x_bis] != -2 and C_l[molecule_x_bis] != -2 and len(assez) == 0:

        M_lf.insert(t2, ro_l[molecule_x_bis] / C_l[molecule_x_bis])

        assez.insert(0, 1)

    if ro_l[molecule_x_bis] != -2 and v_l[molecule_x_bis] != -2 and n_l[molecule_x_bis] != -2 and len(assez) == 0:

        M_lf.insert(t2, (ro_l[molecule_x_bis] * v_l[molecule_x_bis]) / n_l[molecule_x_bis])

        assez.insert(0, 1)

def v_fun():

    t2 = len(v_lf) + 1

    if v_l[molecule_x_bis] != -2 and len(assez) == 0:

        v_lf.insert(t2, v_l[molecule_x_bis])

        assez.insert(0, 1)

    if C_l[molecule_x_bis] != -2 and n_l[molecule_x_bis] != -2 and len(assez) == 0:

        v_lf.insert(t2, n_l[molecule_x_bis] / C_l[molecule_x_bis])

        assez.insert(0, 1)

    if m_l[molecule_x_bis] != -2 and ro_l[molecule_x_bis] != -2 and len(assez) == 0:

        v_lf.insert(t2, m_l[molecule_x_bis] / ro_l[molecule_x_bis])

        assez.insert(0, 1)

    if n_l[molecule_x_bis] != -2 and M_l[molecule_x_bis] != -2 and ro_l[molecule_x_bis] != -2 and len(assez) == 0:

        v_lf.insert(t2, (n_l[molecule_x_bis] * M_l[molecule_x_bis]) / ro_l[molecule_x_bis])

        assez.insert(0, 1)

    if n_l[molecule_x_bis] != -2 and len(r_equa_bis) > 0 and ro_l[molecule_x_bis] != -2 and len(assez) == 0:

        v_lf.insert(t2, (n_l[molecule_x_bis] * r_equa_bis_val[molecule_x_bis2]) / ro_l[molecule_x_bis])

        assez.insert(0, 1)

def C_fun():

    t2 = len(C_lf) + 1

    if C_l[molecule_x_bis] != -2 and len(assez) == 0:

        C_lf.insert(t2, C_l[molecule_x_bis])

        assez.insert(0, 1)

    if n_l[molecule_x_bis] != -2 and v_l[molecule_x_bis] != -2 and len(assez) == 0:

        C_lf.insert(t2, n_l[molecule_x_bis] / v_l[molecule_x_bis])

        assez.insert(0, 1)

    if ro_l[molecule_x_bis] != -2 and M_l[molecule_x_bis] != -2 and len(assez) == 0:

        C_lf.insert(t2, ro_l[molecule_x_bis] / M_l[molecule_x_bis])

        assez.insert(0, 1)

    if ro_l[molecule_x_bis] != -2 and len(r_equa_bis) > 0 and len(assez) == 0:

        C_lf.insert(t2, ro_l[molecule_x_bis] / r_equa_bis_val[molecule_x_bis2])

        assez.insert(0, 1)

def m_fun():

    t2 = len(n_lf) + 1

    if m_l[molecule_x_bis] != -2 and len(assez) == 0:

        m_lf.insert(t2, m_l[molecule_x_bis])

        assez.insert(0, 1)

    if M_l[molecule_x_bis] != -2 and n_l[molecule_x_bis] != -2 and len(assez) == 0:

        m_lf.insert(t2, M_l[molecule_x_bis] * n_l[molecule_x_bis])

        assez.insert(0, 1)

    if len(r_equa_bis) > 0 and n_l[molecule_x_bis] != -2 and len(assez) == 0:

        m_lf.insert(t2, r_equa_bis_val[molecule_x_bis2] * n_l[molecule_x_bis])

        assez.insert(0, 1)

    if M_l[molecule_x_bis] != -2 and C_l[molecule_x_bis] != -2 and v_l[molecule_x_bis] != -2 and len(assez) == 0:

        m_lf.insert(t2, M_l[molecule_x_bis] * C_l[molecule_x_bis] * v_l[molecule_x_bis])

        assez.insert(0, 1)

    if len(r_equa_bis) > 0 and C_l[molecule_x_bis] != -2 and v_l[molecule_x_bis] != -2 and len(assez) == 0:

        m_lf.insert(t2, r_equa_bis_val[molecule_x_bis2] * C_l[molecule_x_bis] * v_l[molecule_x_bis])

        assez.insert(0, 1)

treg = 0

tregb = 0

treg2 = 0

tmax = 0

tmin = 0

tminl = []

tmaxl = []

nbprl = []

nbprlf = []

while molecule_x_bis < len(list_des_inter):

    tmin = list_des_inter[molecule_x_bis][0]

    tminl.append(tmin)

    tmax = list_des_inter[molecule_x_bis][-1] 

    tmaxl.append(tmax)

    molecule_x_bis += 1

tmin = min(tminl)

tmax = max(tmaxl)

print("tmin:", tmin, "tmax:", tmax)

molecule_x_bis = 0

while molecule_x_bis < len(list_des_inter):

    while treg2 <= list_des_inter[molecule_x_bis].count(-1):

        while list_des_inter[molecule_x_bis][treg] != -1 and treg + 1 < len(list_des_inter[molecule_x_bis]):

            if treg + 2 == len(list_des_inter[molecule_x_bis]):

                tregb += 1

            tregb += 1

            treg += 1

        nbprl.append(tregb)

        treg += 1

        tregb = 0

        treg2 += 1

    nbprlf.append(nbprl)

    nbprl = []

    treg = 0

    treg2 = 0

    molecule_x_bis += 1

molecule_x_bis = 0

treg = 0

treg2 = 0

tregb = 0

ltp = []

maxparal = []

print(nbprlf)

while treg < len(nbprlf[0]):

    while treg2 < len(list_des_inter):

        ltp.append(nbprlf[treg2][tregb])

        treg2 += 1

    treg2 = 0

    tregb += 1

    maxparal.append(max(ltp))

    ltp = []

    treg += 1

treg = 0

treg2 = 0

tregb = 0

treg2b = 0

nmin = 0

nmax = 0

subpl = []

l = [inter_Ml, inter_rol, inter_nl, inter_ml, inter_vl, inter_Cl, inter_Qrl] #editable

lb = [[], [], [], [], [], [], []] #editable

nl = [M_l, ro_l, n_l, m_l, v_l, C_l, Qr_l] #editable

nl2 = [[], [], [], [], [], [], []] #editable

t = 0

t2 = 0

while t < len(l):

    while t2 < len(l[t]):

        lb[t].append(l[t][t2])

        t2 += 1

    t2 = 0

    t += 1

lf = [] 

lf2 = []

t = 0

if len(l) > 0:

    lon = len(l[0])

while t < len(l):

    if t > 0 and len(l[t - 1]) < len(l[t]):

        lon = len(l[t])

    t += 1

t = 0

t2 = 0

while t < len(l):

    if len(l[t]) < lon:

        mid = (l[t][0] + l[t][1]) / 2

        l[t].insert(1, mid)

        l[t].insert(2, mid)

        t = 0

    t += 1

t = 0

while t < len(l):

    while t2 < len(l[t]):

        lf.append(l[t][t2])

        t2 += 1

    t2 = 0

    t += 1

t = 0

lonf = len(lf)

while t < lonf:

    lf2.append(min(lf))

    lf.remove(min(lf))

    t += 1

t = 0

while t < len(lf2):

    if lf2[t] == lf2[t - 1] == lf2[t - 2] and t > 1:

        lf2.remove(lf2[t])

        t = 0

    t += 1

t = 0

t2 = 1

t3 = 0

t4 = 0

while t < len(lb):

    while t3 < len(lf2):

        if t2 < len(lb[t]):

            if lb[t][t2] < lf2[t3]:

                t2 += 2

                t4 += 1

        if t4 < len(nl[t]):

            nl2[t].append(nl[t][t4])

        t3 += 1

    t2 = 1

    t3 = 0

    t4 = 0

    t += 1

print(nl2)

while molecule_x_bis < len(m_l) - (len(reactifs_f) + len(produits_f) - 2): 

    assez = []

    n_fun()

    if len(assez) == 0:

        n_lf.insert(len(n_lf) + 1, -2)

    assez = []

    ro_fun()

    if len(assez) == 0:

        ro_lf.insert(len(ro_lf) + 1, -2)

    assez = []

    M_fun()

    if len(assez) == 0:

        M_lf.insert(len(M_lf) + 1, -2)

    assez = []

    v_fun()

    if len(assez) == 0:

        v_lf.insert(len(v_lf) + 1, -2)

    assez = []

    C_fun()

    if len(assez) == 0:

        C_lf.insert(len(C_lf) + 1, -2)

    assez = []

    m_fun()

    if len(assez) == 0:

        m_lf.insert(len(m_lf) + 1, -2)

    molecule_x_bis += 1

t_inter = t_inter + 1

molecule_x_bis = 0

molecule_x_bis3 = 0

molecule_x_bix2 = 0

t_inter = 0

arret = 0

assez = []

T4 = 0

print("n_lf:", n_lf)

print("m_lf:", m_lf)

print("M_lf:", M_lf)

print("ro_lf", ro_lf)

print("C_lf", C_lf)

print("v_lf", v_lf)

print(inter_Ml)

print(inter_rol, inter_nl)

formula_lf = []

a = sheet.cell(row=9, column=3).value

a = " ".join(a)

split = str.split(a)

temp_nb = []

prioridad_l = ["*", "raccord", "*", -1, "exp", "log", "ln", -1, "*", "/", -1, "+", "-"]

ressources = ["[", "]", "t", "v", "m", "M", "ro", "+", "-", "(", ")", "*", "/", "**", "exp", "ln", "log"]

str_nb = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0", "0.3"]

formula = []

formula_f = []

t = 0

C = 3

resultat = []

para = 0

position = 0

t2 = 0

nb_potential = 0

cro_count = 0

u2 = 0

index_l = []

t_plusdinspi = 0

while t < len(split):

    u = 0

    resultat.append(0)

    index_l.append(0)

    if split[t] == "r":

        formula.append("r")

    if split[t] == "|":

        formula.append("|")

    if split[t] == "!":

        formula.append("!")

    if split[t] == "e":

        formula.append("exp")

    if split[t] == "l" and split[t + 1] == "o" and split[t + 2] == "g":
        formula.append("log")

    if split[t] == "l" and split[t + 1] == "n":

        formula.append("ln")

    if split[t] == "s" and split[t + 1] == "i" and split[t - 1] != "a":

        formula.append("sin")

    if split[t] == "t" and split[t + 1] == "a" and split[t - 1] != "a":

        formula.append("tan")

    if split[t] == "c" and split[t + 1] == "o" and split[t - 1] != "a":

        formula.append("cos")

    if split[t] == "a" and split[t + 1] == "s":

        formula.append("asin")

    if split[t] == "a" and split[t + 1] == "c":

        formula.append("acos")

    if split[t] == "a" and split[t + 1] == "t":

        formula.append("atan")

    if split[t] == "C" and split[t + 1] == "_":

        formula.append("C")

        split.insert(t + 2, "0.3")

        t_plusdinspi = t + 2

        laterriblefin = 0

        passela = 0

        while laterriblefin == 0:   

            if str_nb.count(split[t_plusdinspi]) > 0:

                if split[t_plusdinspi] != "0.3":

                    lajoutstr = split[t_plusdinspi]

                    lajoutstr2 = formula[-1]

                    lajoutstr3 = lajoutstr2 + lajoutstr

                    formula[-1] = lajoutstr3

            if t_plusdinspi + 2 > len(split):

                laterriblefin = 1

            if str_nb.count(split[t_plusdinspi]) == 0:

                laterriblefin = 1

            t_plusdinspi += 1

        if laterriblefin == 1:

            t_plusdinspi = t + 2

            while str_nb.count(split[t_plusdinspi]) > 0 and t_plusdinspi + 1 < len(split):

                split.pop(t_plusdinspi)

            if t_plusdinspi + 1 == len(split) or str_nb.count(split[t_plusdinspi]) > 0:

                split.pop(t_plusdinspi - 1)

                if str_nb.count(split[t_plusdinspi - 1]) > 0:

                    split.pop(t_plusdinspi - 1)

    if split[t] == "m" and split[t + 1] == "_":

        formula.append("m")

        split.insert(t + 2, "0.3")

        t_plusdinspi = t + 2

        laterriblefin = 0

        passela = 0

        while laterriblefin == 0:   

            if str_nb.count(split[t_plusdinspi]) > 0:

                if split[t_plusdinspi] != "0.3":

                    lajoutstr = split[t_plusdinspi]

                    lajoutstr2 = formula[-1]

                    lajoutstr3 = lajoutstr2 + lajoutstr

                    formula[-1] = lajoutstr3

            if t_plusdinspi + 2 > len(split):

                laterriblefin = 1

            if str_nb.count(split[t_plusdinspi]) == 0:

                laterriblefin = 1

            t_plusdinspi += 1

        if laterriblefin == 1:

            t_plusdinspi = t + 2

            while str_nb.count(split[t_plusdinspi]) > 0 and t_plusdinspi + 1 < len(split):

                split.pop(t_plusdinspi)

            if t_plusdinspi + 1 == len(split) or str_nb.count(split[t_plusdinspi]) > 0:

                split.pop(t_plusdinspi - 1)

                if str_nb.count(split[t_plusdinspi - 1]) > 0:

                    split.pop(t_plusdinspi - 1)

    if split[t] == "v" and split[t + 1] == "_":

        formula.append("v")

        split.insert(t + 2, "0.3")

        t_plusdinspi = t + 2

        laterriblefin = 0

        passela = 0

        while laterriblefin == 0:   

            if str_nb.count(split[t_plusdinspi]) > 0:

                if split[t_plusdinspi] != "0.3":

                    lajoutstr = split[t_plusdinspi]

                    lajoutstr2 = formula[-1]

                    lajoutstr3 = lajoutstr2 + lajoutstr

                    formula[-1] = lajoutstr3

            if t_plusdinspi + 2 > len(split):

                laterriblefin = 1

            if str_nb.count(split[t_plusdinspi]) == 0:

                laterriblefin = 1

            t_plusdinspi += 1

        if laterriblefin == 1:

            t_plusdinspi = t + 2

            while str_nb.count(split[t_plusdinspi]) > 0 and t_plusdinspi + 1 < len(split):

                split.pop(t_plusdinspi)

            if t_plusdinspi + 1 == len(split) or str_nb.count(split[t_plusdinspi]) > 0:

                split.pop(t_plusdinspi - 1)

                if str_nb.count(split[t_plusdinspi - 1]) > 0:

                    split.pop(t_plusdinspi - 1)

    if split[t] == "t" and split[t + 1] == "_":

        formula.append("t")

        split.insert(t + 2, "0.3")

        t_plusdinspi = t + 2

        laterriblefin = 0

        passela = 0

        while laterriblefin == 0:   

            if str_nb.count(split[t_plusdinspi]) > 0:

                if split[t_plusdinspi] != "0.3":

                    lajoutstr = split[t_plusdinspi]

                    lajoutstr2 = formula[-1]

                    lajoutstr3 = lajoutstr2 + lajoutstr

                    formula[-1] = lajoutstr3

            if t_plusdinspi + 2 > len(split):

                laterriblefin = 1

            if str_nb.count(split[t_plusdinspi]) == 0:

                laterriblefin = 1

            t_plusdinspi += 1

        if laterriblefin == 1:

            t_plusdinspi = t + 2

            while str_nb.count(split[t_plusdinspi]) > 0 and t_plusdinspi + 1 < len(split):

                split.pop(t_plusdinspi)

            if t_plusdinspi + 1 == len(split) or str_nb.count(split[t_plusdinspi]) > 0:

                split.pop(t_plusdinspi - 1)

                if str_nb.count(split[t_plusdinspi - 1]) > 0:

                    split.pop(t_plusdinspi - 1)

    if split[t - 1] == "p" and split[t] == "H" and split[t + 1] == "_":

        formula.append("pH")
        
        split.insert(t + 2, "0.3")

        t_plusdinspi = t + 2

        laterriblefin = 0

        passela = 0

        while laterriblefin == 0:   

            if str_nb.count(split[t_plusdinspi]) > 0:

                if split[t_plusdinspi] != "0.3":

                    lajoutstr = split[t_plusdinspi]

                    lajoutstr2 = formula[-1]

                    lajoutstr3 = lajoutstr2 + lajoutstr

                    formula[-1] = lajoutstr3

            if t_plusdinspi + 2 > len(split):

                laterriblefin = 1

            if str_nb.count(split[t_plusdinspi]) == 0:

                laterriblefin = 1

            t_plusdinspi += 1

        if laterriblefin == 1:

            t_plusdinspi = t + 2

            while str_nb.count(split[t_plusdinspi]) > 0 and t_plusdinspi + 1 < len(split):

                split.pop(t_plusdinspi)

            if t_plusdinspi + 1 == len(split) or str_nb.count(split[t_plusdinspi]) > 0:

                split.pop(t_plusdinspi - 1)

                if str_nb.count(split[t_plusdinspi - 1]) > 0:

                    split.pop(t_plusdinspi - 1)

    if split[t - 1] == "r" and split[t] == "o" and split[t + 1] == "_":

        formula.append("ro")

        split.insert(t + 2, "0.3")

        t_plusdinspi = t + 2

        laterriblefin = 0

        passela = 0

        while laterriblefin == 0:   

            if str_nb.count(split[t_plusdinspi]) > 0:

                if split[t_plusdinspi] != "0.3":

                    lajoutstr = split[t_plusdinspi]

                    lajoutstr2 = formula[-1]

                    lajoutstr3 = lajoutstr2 + lajoutstr

                    formula[-1] = lajoutstr3

            if t_plusdinspi + 2 > len(split):

                laterriblefin = 1

            if str_nb.count(split[t_plusdinspi]) == 0:

                laterriblefin = 1

            t_plusdinspi += 1

        if laterriblefin == 1:

            t_plusdinspi = t + 2

            while str_nb.count(split[t_plusdinspi]) > 0 and t_plusdinspi + 1 < len(split):

                split.pop(t_plusdinspi)

            if t_plusdinspi + 1 == len(split) or str_nb.count(split[t_plusdinspi]) > 0:

                split.pop(t_plusdinspi - 1)

                if str_nb.count(split[t_plusdinspi - 1]) > 0:

                    split.pop(t_plusdinspi - 1)

    if split[t - 1] == "Q" and split[t] == "r" and split[t + 1] == "_":

        formula.append("Qr")
       
        split.insert(t + 2, "0.3")

        t_plusdinspi = t + 2

        laterriblefin = 0

        passela = 0

        while laterriblefin == 0:   

            if str_nb.count(split[t_plusdinspi]) > 0:

                if split[t_plusdinspi] != "0.3":

                    lajoutstr = split[t_plusdinspi]

                    lajoutstr2 = formula[-1]

                    lajoutstr3 = lajoutstr2 + lajoutstr

                    formula[-1] = lajoutstr3

            if t_plusdinspi + 2 > len(split):

                laterriblefin = 1

            if str_nb.count(split[t_plusdinspi]) == 0:

                laterriblefin = 1

            t_plusdinspi += 1

        if laterriblefin == 1:

            t_plusdinspi = t + 2

            while str_nb.count(split[t_plusdinspi]) > 0 and t_plusdinspi + 1 < len(split):

                split.pop(t_plusdinspi)

            if t_plusdinspi + 1 == len(split) or str_nb.count(split[t_plusdinspi]) > 0:

                split.pop(t_plusdinspi - 1)

                if str_nb.count(split[t_plusdinspi - 1]) > 0:

                    split.pop(t_plusdinspi - 1)

    if split[t] == "n" and split[t + 1] == "_":

        formula.append("n")

        split.insert(t + 2, "0.3")

        t_plusdinspi = t + 2

        laterriblefin = 0

        passela = 0

        while laterriblefin == 0:   

            if str_nb.count(split[t_plusdinspi]) > 0:

                if split[t_plusdinspi] != "0.3":

                    lajoutstr = split[t_plusdinspi]

                    lajoutstr2 = formula[-1]

                    lajoutstr3 = lajoutstr2 + lajoutstr

                    formula[-1] = lajoutstr3

            if t_plusdinspi + 2 > len(split):

                laterriblefin = 1

            if str_nb.count(split[t_plusdinspi]) == 0:

                laterriblefin = 1

            t_plusdinspi += 1

        if laterriblefin == 1:

            t_plusdinspi = t + 2

            while str_nb.count(split[t_plusdinspi]) > 0 and t_plusdinspi + 1 < len(split):

                split.pop(t_plusdinspi)

            if t_plusdinspi + 1 == len(split) or str_nb.count(split[t_plusdinspi]) > 0:

                split.pop(t_plusdinspi - 1)

                if str_nb.count(split[t_plusdinspi - 1]) > 0:

                    split.pop(t_plusdinspi - 1)

    if split[t] == "(":

        formula.append(split[t])

    if split[t] == ")":

        formula.append(split[t])

    moved = 0

    while t2 < len(str_nb):

        if split[t] == str_nb[t2] and cro_count != 1:

            position = len(formula)

            temp_nb.append(split[t])

            nb_potential = 1

            t2 = len(str_nb)

            u = 1

            u2 = 1

        if split[t] == "." and cro_count != 1:

            temp_nb.append(split[t])

            t2 = len(str_nb)   

            u = 1

            u2 = 1

        t2 +=  1

    t2 = 0

    if u == 0 and u2 == 1 or u2 == 1 and t + 1 == len(split):

        nb_potential = 0

        temp_nb = "".join(temp_nb)

        formula.insert(position, float(temp_nb))

        temp_nb = []

        u2 = 0

    if split[t] == "+" or split[t] == "-" or split[t] == "*" or split[t] == "/":

        formula.append(split[t])

    t += 1

t = 0

formula2 = []

while t < len(formula):

    formula2.append(formula[t])

    t += 1

t = 0

para = 0

cro_count = 0

ajout_l = []

t2 = 0

t2b = 0

c = formula.count("(")

para_end_count = 0

score_para = []

num_para = []

t3 = 0

t4 = 0

while t3 < len(formula):

    if formula[t3] == "(":

        num_para.insert(para, para)

        para += 1

    t3 += 1

t3 = 0

para_end_count = 0

prem_l = []

len_l = []

while t3 < len(num_para):

    len_l.insert(0, 0)

    score_para.insert(0, 0)

    prem_l.insert(0, 0)

    t3 += 1

t3 = 0

u = 0

t4 = 0

t4 = 0

prem = 0

num_para2 = []

while t4 < len(num_para):

    num_para2.insert(t4, num_para[t4])

    t4 += 1

t4 = 0

para = 0

depart_l = []

fin_l = []

while t3 < len(formula):

    if formula[t3] == "(":

        depart_l.append(t3)

        para += 1

        while t4 < para:

            score_para[t4] += 1

            t4 += 1

        t4 = 0

    if formula[t3] == ")":

        fin_l.append(t3)

        while t4 < para:

            score_para[t4] -= 1

            if score_para[t4] > 0:

                len_l[t4] += 1

            if score_para[t4] == 0 and prem_l[t4] != -2:

                prem_l[t4] = -2

                u = num_para2[t4]

                num_para.insert(num_para.index(u) + len_l[t4] * 2 + 1, u)

            t4 += 1

        t4 = 0

    t3 += 1

t3 = 0

t4 = 0

num_para2 = []

while t4 < len(num_para):

    num_para2.insert(t4, num_para[t4])

    t4 += 1 

t4 = 0

t5 = 0

egal = 0

passed = 0

passed2 = 0

autorise = 0

nouveau_para = -1

t6_f = 0

ctn = 0

t6 = 0

t7 = 0

para_joker = 0

last = 0

t8 = 0

rela_t = 0

last_index = []

long_num_para = len(num_para)

rela_t_f = 0

num_para_persistant = []

t_subliminal = 0

while t_subliminal < len(num_para):

    num_para_persistant.append(num_para[t_subliminal])

    t_subliminal += 1 

t = 0

while t < len(formula):

    if type(formula[t]) == float and formula[t - 1] == "-" and type(formula[t - 2]) == str \
            or type(formula[t]) == float and formula[t - 1] == "-" and t == 1:

        formula[t - 1] = -formula[t]

        formula.pop(t)

    t += 1

t = 0

t_tron = 0

c_la = -1

while t_tron < len(formula):

    if formula[t_tron] == "(":

        c_la += 1

    if t_tron + 2 < len(formula):

        if formula[t_tron] == "(" and type(formula[t_tron + 1]) == float and formula[t_tron + 2] == ")":

            formula[t_tron] = formula[t_tron + 1]

            formula.pop(t_tron + 1)

            formula.pop(t_tron + 1)

            num_para.remove(c_la)

            num_para.remove(c_la)

            num_para2.remove(c_la)

            num_para2.remove(c_la)

    t_tron += 1

t_tron = 0

moved = 0

t2 = 0

position_l = []

position = 0

while t2 < len(formula):

    if formula[t2] == "(":

        position_l.append(t2)

        position += 1

    if formula[t2] == ")":

        position_l.append(t2)

    t2 += 1

t2 = 0

para = 0

complete = 0

depart = 0

t = 0

t_num_para = 0

sommus_l = []

t_sommus = 0

while t < len(formula):

    sommus_l.append(0)

    t += 1

t = 0

moved = 0

para_ferm = 0

beg = 0

deca = 0

sommus_val = 0

t = 0

prc = 0

t_dvr = 0

formula_lf = []

t_fact = 0

facto = 0

while len(num_para) > 0:

    t = para + sommus_val

    while formula[t] != ")":   

        t += 1

    para_ferm = t 

    if beg == 1:

        formula.pop(para + sommus_val)   

        formula.pop(para_ferm - 1)

        sommus_l[para] = sommus_l[para] - 1
    
        sommus_l[para + (para_ferm - para - sommus_val) - 1] = sommus_l[para + (para_ferm - para - sommus_val) - 1] - 1
        
        num_para.remove(deca)

        num_para.remove(deca)

    eldmn = 0

    t = 0

    beg = 1

    while t < len(num_para):

        if num_para[t - 1] == num_para[t] and t > 0:

            deca = num_para[t]

            para = position_l[num_para_persistant.index(deca)]

            t = len(num_para)

        t += 1

    sommus_val = 0

    while t_sommus < para:

        sommus_val = sommus_val + sommus_l[t_sommus]

        t_sommus += 1

    t_sommus = 0

    prc = 0

    t = para + sommus_val

    tdeff = para + sommus_val

    while formula[t] != ")" and len(num_para) > 0:

        if type(formula[t - 1]) == float and formula[t] == "!":

            formula[t - 1] = int(formula[t - 1])

            formula[t - 1] = math.factorial(formula[t - 1])

            formula[t - 1] = float(formula[t - 1])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1   

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "r" and type(formula[t]) == float:

            formula[t - 1] = math.radians(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1   

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "ln" and type(formula[t]) == float:

            formula[t - 1] = math.log(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1   

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "exp" and type(formula[t]) == float:

            formula[t - 1] = math.exp(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "log" and type(formula[t]) == float:

            formula[t - 1] = math.log(formula[t]) / math.log(10)

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "sin" and type(formula[t]) == float:

            formula[t - 1] = math.sin(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "tan" and type(formula[t]) == float:

            formula[t - 1] = math.tan(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val           

        if formula[t - 1] == "cos" and type(formula[t]) == float:

            formula[t - 1] = math.cos(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "sqr" and type(formula[t]) == float:

            formula[t - 1] = math.sqrt(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val 

        if formula[t - 1] == "acos" and type(formula[t]) == float:

            formula[t - 1] = math.acos(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "asin" and type(formula[t]) == float:

            formula[t - 1] = math.asin(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "atan" and type(formula[t]) == float:

            formula[t - 1] = math.atan(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "atanh" and type(formula[t]) == float:

            formula[t - 1] = math.atanh(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val      

        if formula[t - 1] == "asinh" and type(formula[t]) == float:

            formula[t - 1] = math.asinh(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "acosh" and type(formula[t]) == float:

            formula[t - 1] = math.acosh(formula[t])

            formula.pop(t)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 1

            moved = 1

            t = para + sommus_val

        t += 1

    t = para + sommus_val 

    while formula[t] != ")" and len(num_para) > 0:

        if formula[t - 1] == "*" and formula[t - 2] == "*" and type(formula[t]) == float\
                and type(formula[t - 3]) == float:

            formula[t - 3] = formula[t - 3] ** formula[t]

            formula.pop(t - 2)

            formula.pop(t - 2)

            formula.pop(t - 2)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 3

            moved = 1

            t = para + sommus_val

        t += 1

    t = para + sommus_val

    tdeff = para + sommus_val

    while formula[t] != ")" and len(num_para) > 0:

        if formula[t - 1] == "*" and type(formula[t - 2]) == float and type(formula[t]) == float:

            formula[t - 2] = formula[t - 2] * formula[t]

            formula.pop(t - 1)

            formula.pop(t - 1)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 2

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "/" and type(formula[t - 2]) == float and type(formula[t]) == float:

            formula[t - 2] = formula[t - 2] / formula[t]

            formula.pop(t - 1)

            formula.pop(t - 1)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 2

            moved = 1

            t = para + sommus_val

        t += 1

    t = para + sommus_val

    while formula[t] != ")" and len(num_para) > 0:

        if formula[t - 1] == "+" and type(formula[t - 2]) == float and type(formula[t]) == float:

            formula[t - 2] = formula[t - 2] + formula[t]

            formula.pop(t - 1)

            formula.pop(t - 1)

            sommus_l[para + abs(t - tdeff)] = sommus_l[para + abs(t - tdeff)] - 2

            moved = 1

            t = para + sommus_val

        if formula[t - 1] == "-" and type(formula[t - 2]) == float and type(formula[t]) == float:

            formula[t - 2] = formula[t - 2] - formula[t]

            formula.pop(t - 1)

            formula.pop(t - 1)

            sommus_l[para + abs(tdeff - t)] = sommus_l[para + abs(t - tdeff)] - 2

            moved = 1

            t = para + sommus_val

        t += 1

    t = para + sommus_val

    while formula[t] != ")" and len(num_para) > 0:

        if formula[t - 2] == "|" and formula[t] == "|" and type(formula[t - 1]) == float:
            
            formula[t - 2] = abs(formula[t - 1])

            formula.pop(t - 1)

            formula.pop(t - 1)

            sommus_l[para + abs(tdeff - t)] = sommus_l[para + abs(t - tdeff)] - 2

            moved = 1

            t = para + sommus_val

        t += 1

t = 0

while t < len(formula):

    if formula[t - 1] == "r" and type(formula[t]) == float:

        formula[t - 1] = math.radians(formula[t])

        formula.pop(t)

        t = 0

        moved = 0
       
    if type(formula[t - 1]) == float and formula[t] == "!":

        formula[t - 1] = int(formula[t - 1])

        formula[t - 1] = math.factorial(formula[t - 1])

        formula[t - 1] = float(formula[t - 1])

        formula.pop(t)

        t = 0

    if formula[t - 1] == "ln" and type(formula[t]) == float:

        formula[t - 1] = math.log(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "exp" and type(formula[t]) == float:

        formula[t - 1] = math.exp(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "log" and type(formula[t]) == float:

        formula[t - 1] = math.log(formula[t]) / math.log(10)

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "sin" and type(formula[t]) == float:

        formula[t - 1] = math.sin(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "tan" and type(formula[t]) == float:

        formula[t - 1] = math.tan(formula[t])

        formula.pop(t)

        moved = 1

        t = 0       

    if formula[t - 1] == "cos" and type(formula[t]) == float:

        formula[t - 1] = math.cos(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "sqr" and type(formula[t]) == float:

        formula[t - 1] = math.sqrt(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "acos" and type(formula[t]) == float:

        formula[t - 1] = math.acos(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "asin" and type(formula[t]) == float:

        formula[t - 1] = math.asin(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "atan" and type(formula[t]) == float:

        formula[t - 1] = math.atan(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "atanh" and type(formula[t]) == float:

        formula[t - 1] = math.atanh(formula[t])

        formula.pop(t)

        moved = 1

        t = 0 

    if formula[t - 1] == "asinh" and type(formula[t]) == float:

        formula[t - 1] = math.asinh(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "acosh" and type(formula[t]) == float:

        formula[t - 1] = math.acosh(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    t += 1

t = 0

while t < len(formula):

    if formula[t - 1] == "*" and formula[t - 2] == "*" and type(formula[t]) == float\
            and type(formula[t - 3]) == float:

        formula[t - 3] = formula[t - 3] ** formula[t]

        formula.pop(t - 2)

        formula.pop(t - 2)

        formula.pop(t - 2)

        moved = 1

        t = 0

    t += 1

t = 0

tdeff = para + sommus_val

while t < len(formula):

    if formula[t - 1] == "*" and type(formula[t - 2]) == float and type(formula[t]) == float:

        formula[t - 2] = formula[t - 2] * formula[t]

        formula.pop(t - 1)

        formula.pop(t - 1)

        moved = 1

        t = 0

    if formula[t - 1] == "/" and type(formula[t - 2]) == float and type(formula[t]) == float:

        formula[t - 2] = formula[t - 2] / formula[t]

        formula.pop(t - 1)

        formula.pop(t - 1)

        moved = 1

        t = 0

    t += 1

t = 0

while t < len(formula):

    if formula[t - 1] == "+" and type(formula[t - 2]) == float and type(formula[t]) == float:

        formula[t - 2] = formula[t - 2] + formula[t]

        formula.pop(t - 1)

        formula.pop(t - 1)

        moved = 1

        t = 0

    if formula[t - 1] == "-" and type(formula[t - 2]) == float and type(formula[t]) == float:

        formula[t - 2] = formula[t - 2] - formula[t]

        formula.pop(t - 1)

        formula.pop(t - 1)

        moved = 1

        t = 0

    t += 1

t = 0

while t < len(formula):

    if formula[t - 1] == "|" and type(formula[t]) == float:

        formula[t - 1] = abs(formula[t])

        formula.pop(t)

        formula.pop(t)

        moved = 1

    t += 1

t = 0

while t < len(formula):

    if formula[t - 1] == "r" and type(formula[t]) == float:

        formula[t - 1] = math.radians(formula[t])

        formula.pop(t)

        t = 0

        moved = 0
       
    if type(formula[t - 1]) == float and formula[t] == "!":

        formula[t - 1] = int(formula[t - 1])

        formula[t - 1] = math.factorial(formula[t - 1])

        formula[t - 1] = float(formula[t - 1])

        formula.pop(t)

        t = 0

    if formula[t - 1] == "ln" and type(formula[t]) == float:

        formula[t - 1] = math.log(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "exp" and type(formula[t]) == float:

        formula[t - 1] = math.exp(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "log" and type(formula[t]) == float:

        formula[t - 1] = math.log(formula[t]) / math.log(10)

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "sin" and type(formula[t]) == float:

        formula[t - 1] = math.sin(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "tan" and type(formula[t]) == float:

        formula[t - 1] = math.tan(formula[t])

        formula.pop(t)

        moved = 1

        t = 0       

    if formula[t - 1] == "cos" and type(formula[t]) == float:

        formula[t - 1] = math.cos(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "sqr" and type(formula[t]) == float:

        formula[t - 1] = math.sqrt(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "acos" and type(formula[t]) == float:

        formula[t - 1] = math.acos(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "asin" and type(formula[t]) == float:

        formula[t - 1] = math.asin(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "atan" and type(formula[t]) == float:

        formula[t - 1] = math.atan(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "atanh" and type(formula[t]) == float:

        formula[t - 1] = math.atanh(formula[t])

        formula.pop(t)

        moved = 1

        t = 0 

    if formula[t - 1] == "asinh" and type(formula[t]) == float:

        formula[t - 1] = math.asinh(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    if formula[t - 1] == "acosh" and type(formula[t]) == float:

        formula[t - 1] = math.acosh(formula[t])

        formula.pop(t)

        moved = 1

        t = 0

    t += 1

t = 0

while t < len(formula):

    if formula[t - 1] == "*" and formula[t - 2] == "*" and type(formula[t]) == float\
            and type(formula[t - 3]) == float:

        formula[t - 3] = formula[t - 3] ** formula[t]

        formula.pop(t - 2)

        formula.pop(t - 2)

        formula.pop(t - 2)

        moved = 1

        t = 0

    t += 1

t = 0

tdeff = para + sommus_val

while t < len(formula):

    if formula[t - 1] == "*" and type(formula[t - 2]) == float and type(formula[t]) == float:

        formula[t - 2] = formula[t - 2] * formula[t]

        formula.pop(t - 1)

        formula.pop(t - 1)

        moved = 1

        t = 0

    if formula[t - 1] == "/" and type(formula[t - 2]) == float and type(formula[t]) == float:

        formula[t - 2] = formula[t - 2] / formula[t]

        formula.pop(t - 1)

        formula.pop(t - 1)

        moved = 1

        t = 0

    t += 1

t = 0

while t < len(formula):

    if formula[t - 1] == "+" and type(formula[t - 2]) == float and type(formula[t]) == float:

        formula[t - 2] = formula[t - 2] + formula[t]

        formula.pop(t - 1)

        formula.pop(t - 1)

        moved = 1

        t = 0

    if formula[t - 1] == "-" and type(formula[t - 2]) == float and type(formula[t]) == float:

        formula[t - 2] = formula[t - 2] - formula[t]

        formula.pop(t - 1)

        formula.pop(t - 1)

        moved = 1

        t = 0

    t += 1

print(formula)

formula_lf.append(formula)

t_qr = 0

t_qr2 = 0

btqr = 0

tpqr = 0

while t_qr < inter_nl.count(-1) + 1:

    while inter_nl[t_qr2] != -1:

        if t_qr2 == 0:

            t_qr2 = t_qr

            btqr = inter_nl[t_qr]

        if inter_nl[t_qr2] > inter_nl[t_qr2 - 1]:

            tpqr = inter_nl[t_qr + 1]

        t_qr2 += 1

    t_qr2 = 0

    t_qr += 1

t_qr = 0

print("bottom", btqr)

print("top",tpqr)


