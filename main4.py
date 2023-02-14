from collections import Counter

from tkinter import *

import time

import matplotlib.pyplot as plt

from mpl_toolkits.mplot3d import Axes3D

import sys

from math import *

from openpyxl import load_workbook

from inspect import currentframe, getframeinfo

import time

import math

import statistics

import numpy as np

cf = currentframe()

w2 = load_workbook("output.xlsx")

sheet2 = w2.active

w = load_workbook("infos.xlsx")

sheet = w.active

coeff_du_requa = []

formule_chem = "C2H6+N2" #can be edited to whatever chemical you want

r_equa = []

r_equa2 = []

t = 0

T = 0

T2 = 0

reactifs = []

reactifs2 = []

r2 = str

reactifs_f = []

passe = 0

produits = []

produits2 = []

p1 = str

produits_f = []

if len(formule_chem) > 1:

    formule_chem = " ".join(formule_chem)

    spll = str.split(formule_chem)

    while t < len(spll):

        if passe == 0:
            reactifs.insert(t, spll[t])

        if passe == 1:
            produits.insert(t, spll[t])

        if spll[t] == "=":
            passe = 1

        t = t + 1

    t = 0

    while t < len(reactifs):

        r2 = "".join(reactifs2)

        if reactifs[t] == "+" or reactifs[t] == "-" or t + 1 == len(reactifs):

            reactifs_f.insert(t, r2)

            reactifs2 = []

        else:

            reactifs2.insert(t, reactifs[t])

        t = t + 1

    t = 0

    while t < len(produits):

        p2 = "".join(produits2)

        if produits[t] == "+" or produits[t] == "-" or t + 1 == len(produits):

            if t + 1 == len(produits):

                produits2.insert(t, produits[t])

            p2 = "".join(produits2)

            produits_f.insert(t, p2)

            produits2 = []

        else:

            produits2.insert(t, produits[t])

        t = t + 1

    t = 0

    while T < len(reactifs_f) + len(produits_f):

        name_list = ["H", "He", "Li", "Be", "B", "C", "N", "O", "F", "Ne", "Na", "Mg", "Al", "Si", "P", "S", "Cl", "Ar",
                     "K"
            , "Ca", "Sc", "Ti", "V", "Cr", "Mn", "Fe", "Co", "Ni", "Cu", "Zn", "Ga", "Ge", "As", "Se", "Br", "Kr", "Rb",
                     "Sr",
                     "Y", "Zr", "Nb", "Mo", "Tc", "Ru", "Rh", "Pd", "Ag", "Cd", "In", "Sn", "Sb", "Te", "I", "Xe", "Cs",
                     "Ba", "Hf"
            , "Ta", "W", "Re", "Os", "Ir", "Pt", "Au", "Hg", "Ti", "Pb", "Bi", "Po", "At", "Rn", "Fr", "Ra", "Rf", "Db",
                     "Sg", "Bh", "Hs", "Mt", "Ds", "Rg", "Cn", "Nh", "Fl", "Mc", "Lv", "Ts", "Og", "La", "Ce", "Pr",
                     "Nd",
                     "Pm",
                     "Sm", "Eu", "Gd", "Tb", "Dy", "Ho", "Er", "Tm", "Yb", "Lu", "Ac", "Th", "Pa", "U", "Np", "Pu",
                     "Am",
                     "Cm",
                     "Bk", "Cf", "Es", "Fm", "Md", "No", "Lr"]

        M_list = [1, 4, 7, 9, 11, 12, 14, 16, 19, 20, 23, 24, 27, 28, 31, 32, 35.5, 40, 39, 40, 45, 48, 51, 52, 55, 56
            , 59, 59, 64, 65, 70, 73, 75, 79, 80, 84, 85, 88, 89, 91, 93, 96, 98, 101, 103, 106, 108, 112, 115, 119,
                  122,
                  128, 127, 131, 133, 137, 178, 181, 184, 186, 190, 192, 195, 197, 201, 204, 207, 209, 209, 210, 222,
                  223,
                  226, 267, 268, 269, 270, 269, 278, 281, 281, 285, 286, 289, 289, 293, 294, 294, 139, 140, 141, 144,
                  145,
                  150, 152, 157, 159, 163, 165, 167, 169, 173, 175, 227, 232, 231, 238, 237, 244, 243, 247, 247, 251,
                  252,
                  257, 258, 259, 262]

        name_numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]

        value_numbers = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]

        crochet_name1 = ["(", "["]

        crochet_name2 = [")", "]"]

        par_cro_mult_list = [0]

        par_cro_numb2 = 0

        par_cro_mult_numb = -1

        para_cro_coeff_list = []

        ti = 0

        ti2 = 0

        t_crochet2 = 0

        t_crochet2bis = 0

        t_crochet2bis2 = 0

        t_crochet2bis2_2 = 0

        t_crochet2_list = []

        fois_list = []

        fois_list2 = []

        t_fois2 = 0

        t_crochet = 0

        para_cro_passage = 0

        M_list_global = []

        t_global = 0

        M_list_global2 = []

        t_list_global = 0

        par_cro_i = 0

        par_cro_i2 = 1

        pp_element2 = 0

        suivant = 0

        zero = 0

        coeff_mult = 0

        coeff_stoechio = 0

        mult_coeff_stoechio = 1

        coeff_stoechio_passage = 0

        coeff_mult_passage = 1

        par_cro_mult_list_second = []

        t_crochet2bis3 = 0

        t_fois = 0

        M_element_chem = 0

        M_element_chem2 = 0

        M_coeff_mult = 0

        if T < len(reactifs_f):

            molecule = reactifs_f[T]

        else:

            molecule = produits_f[T2]

            T2 = T2 + 1

        molecule = " ".join(molecule)

        spl = str.split(molecule)

        spl2 = []

        t_spl = 0

        passagelist1 = []

        passage_numb = 0

        t4_spl = 0

        t3_spl = 0

        t_f = 0

        t5bis = 1

        po = -1

        t_fbis = 0

        t_flist = []

        tpassage = 0

        passage_entre_list = []

        list_eliminate = []

        il_y_a_une_para = 0

        while t_spl < len(spl):

            if spl[t_spl] == "(" or spl[t_spl] == "[":

                list_eliminate.insert(0, 1)

                passage_entre_list.insert(0, 0)

                t5bis = 1

                po = -1

                spl2.insert(len(spl) + 1, t_f)

                passagelist1.insert(t_f, 0)

                t_fbis = 0

                t_flist.insert(t_f, t_f)

                t_f = t_f + 1

                tpassage = tpassage + 1

                while t3_spl < len(passagelist1):

                    passagelist1[t3_spl] = passagelist1[t3_spl] + 1

                    t3_spl = t3_spl + 1

                t3_spl = 0

            else:

                il_y_a_une_para = 1

            if spl[t_spl] == ")" or spl[t_spl] == "]":

                while t3_spl < len(passagelist1):

                    passagelist1[t3_spl] = passagelist1[t3_spl] - 1

                    if passagelist1[t3_spl] == 0:

                        passagelist1.pop()

                        while t4_spl < len(spl2):

                            if spl2[t4_spl] == t_flist[t3_spl]:
                                passage_numb = passage_numb + 1

                            t4_spl = t4_spl + 1

                        if passage_numb > 1:

                            spl2.insert(len(spl) + 1, t_flist[-t5bis])

                            t5bis = t5bis + 1

                            passage_numb = 0

                            t5 = 0

                        else:

                            po = -1

                            t5bis = 1

                            spl2.insert(len(spl) + 1, t_flist[t3_spl])

                            passage_numb = 0

                        t4_spl = 0

                        t_fbis = t_fbis + 1

                    t3_spl = t3_spl + 1

                t3_spl = 0

            else:

                if il_y_a_une_para == 1:

                    il_y_a_une_para = 2

            if il_y_a_une_para == 2:

                spl2.insert(t_spl, spl[t_spl])

            il_y_a_une_para = 0

            t_spl = t_spl + 1

        t_spl = 0

        while t_spl < len(spl2):

            if type(spl2[t_spl]) == int:

                spl2[t_spl] = spl2[t_spl] + 1

            t_spl = t_spl + 1

        l = spl2

        ti = 0

        ti2 = 0

        l22 = []

        l33 = []

        l33b = []

        l4 = []

        l3 = []

        l5 = []

        while ti < len(l):

            if type(l[ti]) != str:

                l22.insert(ti2, l[ti])

                ti2 = ti2 + 1

            ti = ti + 1

        ti = 0

        while ti < len(l22) / 2 + 1:

            l33.insert(ti, ti)

            ti = ti + 1

        l33.remove(l33[0])

        l4 = l33

        l2 = []

        t = 0

        t2 = 0

        t3 = 0

        t4 = 0

        while t < len(l4):

            l3.insert(0, 1)

            l5.insert(0, 0)

            t = t + 1

        t = 0

        t5 = 0

        while t < len(l):

            while t2 < t3:

                l2.insert(t, l[t])

                if l2[t2] == l[t2] and type(l2[t2]) != str:

                    l3[l[t2] - 1] = 0

                t2 = t2 + 1

            t2 = 0

            l2 = []

            while t5 < t4:

                if l3[t5] != 0:

                    l5[t5] = l5[t5] + 1

                t5 = t5 + 1

            t5 = 0

            if t4 < len(l4):

                if l[t] == l4[t4]:

                    t4 = t4 + 1

            t = t + 1

            t3 = t3 + 1

        spl_final = []

        M_passé = 0

        t2 = 0

        l = []

        t_entre = 0

        list_entre = []

        pp_element = 0

        trouvé = 0

        min_list = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q", "r",
                    "s", "t", "u", "v", "w", "x", "y", "z"]
        spl2 = []

        t_spl = 0

        t2_spl = 0

        passagelist1 = []

        passage_numb = 0

        t4_spl = 0

        t5_spl = 0

        t3_spl = 0

        t_f = 0

        t5bis = 1

        po = -1

        t_fbis = 0

        t_flist = []

        tpassage = 0

        while t_spl < len(spl):

            if spl[t_spl] == "(" or spl[t_spl] == "[":

                list_entre.insert(0, 0)

                t5bis = 1

                po = -1

                spl2.insert(len(spl) + 1, t_f)

                passagelist1.insert(t_f, 0)

                t_fbis = 0

                t_flist.insert(t_f, t_f)

                t_f = t_f + 1

                tpassage = tpassage + 1

                while t3_spl < len(passagelist1):
                    passagelist1[t3_spl] = passagelist1[t3_spl] + 1

                    t3_spl = t3_spl + 1

                t3_spl = 0

            if spl[t_spl] == ")" or spl[t_spl] == "]":

                while t3_spl < len(passagelist1):

                    passagelist1[t3_spl] = passagelist1[t3_spl] - 1

                    if passagelist1[t3_spl] == 0:

                        passagelist1.pop()

                        while t4_spl < len(spl2):

                            if spl2[t4_spl] == t_flist[t3_spl]:
                                passage_numb = passage_numb + 1

                            t4_spl = t4_spl + 1

                        if passage_numb > 1:

                            spl2.insert(len(spl) + 1, t_flist[-t5bis])

                            t5bis = t5bis + 1

                            passage_numb = 0

                            t5 = 0

                        else:

                            po = -1

                            t5bis = 1

                            spl2.insert(len(spl) + 1, t_flist[t3_spl])

                            passage_numb = 0

                        t4_spl = 0

                        t_fbis = t_fbis + 1

                    t3_spl = t3_spl + 1

                t3_spl = 0

            t_spl = t_spl + 1

        coco_list = []

        element_de_passage = 0

        t2bis = -1

        while t2 < len(spl):

            while spl[t2] != min_list[t2bis] and t2bis < len(min_list) - 1:

                t2bis = t2bis + 1

                if spl[t2] == min_list[t2bis]:
                    spl_final.pop()

                    l.insert(0, spl[t2 - 1])

                    l.insert(1, min_list[t2bis])

                    l = "".join(l)

                    spl_final.insert(t2, l)

                    l = []

                    trouvé = 1

            if trouvé == 0:
                spl_final.insert(t2, spl[t2])

            t2 = t2 + 1

            t2bis = -1

            trouvé = 0

        t2 = 0

        spl = []

        while t2 < len(spl_final):
            spl.insert(t2, spl_final[t2])

            t2 = t2 + 1

        t = 0

        t_coeff = 0

        t_element_chem = 0

        dernier_M = 0

        M_element_chem_coeff = 0

        t_coco = 0

        while t < len(spl):

            while t_element_chem < len(name_list) and spl[t] != name_list[t_element_chem]:
                t_element_chem = t_element_chem + 1

            if t_element_chem == len(name_list):
                t_element_chem = t_element_chem - 1

            if spl[t] == name_list[t_element_chem]:

                pp_element = 1

                element_de_passage = 1

                para_cro_passage = 0

                coeff_stoechio_passage = 0

                if par_cro_mult_numb > -1:

                    coco_list.insert(t_coco, "d")

                    t_coco = t_coco + 1

                    fois_list2.insert(t_fois2, coeff_mult)

                    t_fois2 = t_fois2 + 1

                    t_crochet2bis2 = t_crochet2bis2 + 1

                    t_crochet2bis3 = t_crochet2bis2

                    if par_cro_numb2 == 0:
                        par_cro_mult_list.insert(t_crochet2bis2, -1)

                        t_crochet2bis2 = t_crochet2bis2 + 1

                        par_cro_mult_list_second.insert(t_crochet2bis3, "fini")

                        t_crochet2bis3 = t_crochet2bis3 + 1

                    par_cro_numb2 = par_cro_numb2 + M_list[t_element_chem]

                    par_cro_mult_list.insert(t_crochet2bis2, par_cro_numb2)

                    par_cro_mult_list_second.insert(t_crochet2bis2, M_list[t_element_chem])

                    M_element_chem2 = M_element_chem2 + M_list[t_element_chem]

                    M_list_global2.insert(t_list_global, M_list[t_element_chem])

                    t_list_global = t_list_global + 1

                    M_list_global.insert(t_list_global, "/")

                    t_list_global = t_list_global + 1

                else:

                    M_element_chem = M_element_chem + M_list[t_element_chem]

                dernier_M = t_element_chem

            t_element_chem = 0

            while t_coeff < len(name_numbers) and spl[t] != name_numbers[t_coeff]:
                t_coeff = t_coeff + 1

            if t_coeff == len(name_numbers):
                t_coeff = t_coeff - 1

                coeff_mult = 0

                coeff_mult_passage = 1

                M_coeff_mult = 0

            if spl[t] == name_numbers[t_coeff]:

                if par_cro_mult_numb > -1 or para_cro_passage == 1:

                    if para_cro_passage == 1:

                        par_cro_i = value_numbers[t_coeff]

                        para_cro_coeff_list.insert(ti, par_cro_i)

                        ti = ti + 1

                    else:

                        coeff_mult = coeff_mult * coeff_mult_passage + value_numbers[t_coeff]

                        if pp_element == 1:
                            coeff_mult_passage = 10

                            pp_element = 0

                            pp_element2 = 1

                        M_passé = M_coeff_mult

                        M_coeff_mult = coeff_mult * M_list[dernier_M]

                        if pp_element2 == 1:

                            M_element_chem2 = M_element_chem2 + M_coeff_mult - M_list[dernier_M]

                            pp_element2 = 0

                        else:

                            M_element_chem2 = M_element_chem2 + M_coeff_mult - M_passé

                        M_list_global.insert(t_list_global, name_list[dernier_M])

                        M_list_global2.insert(t_list_global, M_coeff_mult)

                        t_list_global = t_list_global + 1

                else:

                    if t == 0 or coeff_stoechio_passage == 1:

                        coeff_stoechio = coeff_stoechio * mult_coeff_stoechio + value_numbers[t_coeff]

                        if t == 0:
                            mult_coeff_stoechio = mult_coeff_stoechio * 10

                        coeff_stoechio_passage = 1

                    else:

                        coeff_mult = coeff_mult * coeff_mult_passage + value_numbers[t_coeff]

                        if pp_element == 1:
                            coeff_mult_passage = 10

                            pp_element = 0

                            pp_element2 = 1

                        M_passé = M_coeff_mult

                        M_coeff_mult = coeff_mult * M_list[dernier_M]

                        if pp_element2 == 1:

                            M_element_chem = M_element_chem + M_coeff_mult - M_list[dernier_M]

                            pp_element2 = 0

                        else:

                            M_element_chem = M_element_chem + M_coeff_mult - M_passé

            while t_crochet < len(crochet_name1) and spl[t] != crochet_name1[t_crochet]:
                t_crochet = t_crochet + 1

            if t_crochet == len(crochet_name1):
                t_crochet = 0

            if spl[t] == crochet_name1[t_crochet]:
                element_de_passage = 0

                para_cro_passage = 0

                coeff_stoechio_passage = 0

                par_cro_mult_numb = par_cro_mult_numb + 1

                par_cro_numb2 = 0

            t_crochet = 0

            while t_crochet < len(crochet_name2) and spl[t] != crochet_name2[t_crochet]:
                t_crochet = t_crochet + 1

            if t_crochet == len(crochet_name2):
                t_crochet = 0

            if spl[t] == crochet_name2[t_crochet]:
                element_de_passage = 0

                para_cro_coeff_list.insert(ti, -2)

                ti = ti + 1

                para_cro_passage = 1

                par_cro_mult_numb = par_cro_mult_numb - 1

            t_coeff = 0

            t = t + 1

            t_crochet = 0

        ti = 0

        ti2 = 0

        para_cro_coeff_list2 = []

        while ti < len(para_cro_coeff_list):

            if para_cro_coeff_list[ti] != -2:
                para_cro_coeff_list2.insert(ti, ti)

                ti2 = 1

            ti = ti + 1

        ti = 1

        t = 0

        t_par = 0

        if t_crochet2bis2 > 0:

            while t < 100:

                if t < len(par_cro_mult_list_second):

                    if par_cro_mult_list_second[t] == "fini":
                        t_par = 1

                if t + 1 < len(coco_list):

                    if coco_list[t + 1] != "d" and coco_list[t] == "d":
                        M_element_chem = M_element_chem + par_cro_mult_list_second[t + t_par] * coco_list[t + 1]

                        t_par = 0

                        t = t + 1

                        t_coco = t_coco + 1

                t = t + 1

        t = 0

        para_cro_coeff_list99 = []

        while t < len(para_cro_coeff_list):
            para_cro_coeff_list99.insert(t, para_cro_coeff_list[t])

            t = t + 1

        idx = 0

        idx2 = 0

        para_cro_coeff_list88 = []

        para_cro_coeff_list88_f = []

        lesstw = []

        valid = len(para_cro_coeff_list99)

        t = 0

        ttt = 0

        if len(para_cro_coeff_list) > 1:

            while t < valid:

                if t + 1 < len(para_cro_coeff_list99):

                    if para_cro_coeff_list99[t] == -2 and para_cro_coeff_list99[t + 1] == -2:
                        para_cro_coeff_list.insert(ttt + 1, 1)

                        ttt = ttt + 1

                if t + 1 < len(para_cro_coeff_list99):

                    if para_cro_coeff_list99[t + 1] == -2 and t + 2 == len(para_cro_coeff_list99):
                        para_cro_coeff_list.insert(ttt + 2, 1)

                ttt = ttt + 1

                t = t + 1

            t = 0

            while t < len(para_cro_coeff_list):

                if para_cro_coeff_list[t - 1] == -2:
                    para_cro_coeff_list88.insert(idx, para_cro_coeff_list[t])

                    idx = idx + 1

                para_cro_coeff_list[t] = str(para_cro_coeff_list[t])

                para_cro_coeff_list[t - 1] = str(para_cro_coeff_list[t - 1])

                if len(para_cro_coeff_list[t - 1]) == len(para_cro_coeff_list[t]):
                    para_cro_coeff_list[t] = int(para_cro_coeff_list[t])

                    para_cro_coeff_list88.insert(idx, para_cro_coeff_list[t])

                    idx = idx + 1

                para_cro_coeff_list[t] = int(para_cro_coeff_list[t])

                para_cro_coeff_list[t] = int(para_cro_coeff_list[t])

                if para_cro_coeff_list[t] == -2 and t > 0:

                    idx = 0

                    while idx < len(para_cro_coeff_list88):
                        para_cro_coeff_list88[idx] = str(para_cro_coeff_list88[idx])

                        idx = idx + 1

                    idx = 0

                    joined = "".join(para_cro_coeff_list88)

                    joined = int(joined)

                    para_cro_coeff_list88_f.insert(idx2, joined)

                    idx2 = idx2 + 1

                    para_cro_coeff_list88 = []

                if para_cro_coeff_list[t] == -2:
                    lesstw.insert(0, t)

                t = t + 1

            t = lesstw[0] + 1

            para_cro_coeff_list88 = []

            while t < len(para_cro_coeff_list):
                para_cro_coeff_list[t] = str(para_cro_coeff_list[t])

                para_cro_coeff_list88.insert(t, para_cro_coeff_list[t])

                t = t + 1

            joined = "".join(para_cro_coeff_list88)

            joined = int(joined)

            para_cro_coeff_list88_f.insert(idx2, joined)

        else:

            para_cro_coeff_list88_f = []

            para_cro_coeff_list88_f.insert(0, 1)

        idx = 0

        t = 0

        while t < len(M_list_global):

            if M_list_global[t - 1] == M_list_global[t] and M_list_global[t] != "/":
                M_list_global2[t - 1] = 0

            if t < len(M_list_global) - 1:

                if M_list_global[t] == "/" and M_list_global[t + 1] != "/":
                    M_list_global2[t] = 0

            t = t + 1

        t = 0

        lenghtM = len(M_list_global2)

        while t < lenghtM:

            if 0 in M_list_global2:
                idx = M_list_global2.index(0)

                M_list_global2.remove(M_list_global2[idx])

            t = t + 1

        t = 1

        list_f = []

        add_para = 0

        t2 = 0

        t3 = 0

        par_cro_mult_list_second.insert(len(par_cro_mult_list_second) + 1, "fini")

        while t < len(par_cro_mult_list_second):

            if par_cro_mult_list_second[t] != "fini":

                add_para = add_para + M_list_global2[t3]

                t3 = t3 + 1

                if par_cro_mult_list_second[t + 1] == "fini":
                    list_f.insert(t2, add_para)

            else:

                t2 = t2 + 1

                add_para = 0

            t = t + 1

        ajout_list = []

        list_f3 = []

        pre_final_l = []

        t = 0

        spl2_aa = []

        ajout = 0

        particulier = 0

        vrai_coeff = []

        vrai_coeff2 = []

        spl2_aa2 = []

        while t < len(list_f):
            list_f3.insert(0, "wrong")

            pre_final_l.insert(0, 0)

            t = t + 1

        t = 0

        while t < len(spl2):

            spl2_aa.insert(t, spl2[t])

            counter78 = Counter(spl2_aa)

            if t < len(para_cro_coeff_list88_f):
                vrai_coeff.insert(t, para_cro_coeff_list88_f[t])

            if counter78[spl2_aa[t]] > 1:
                spl2_aa2.insert(t, spl2[t])

                vrai_coeff2.insert(t, spl2_aa[t])

            t = t + 1

        t = 0

        t78 = 0

        passed = 0

        saved_t78 = 0

        while t < len(vrai_coeff):

            while t78 < len(vrai_coeff):

                if t + t78 < len(vrai_coeff) and spl2_aa2[t] > spl2_aa2[t + t78] and t78 > 0:
                    vrai_coeff[t] = vrai_coeff[t] * vrai_coeff[t + t78]

                t78 = t78 + 1

            t78 = 0

            passed = 0

            t = t + 1

        t = 0

        while t < len(list_f):
            pre_final_l[t] = list_f[vrai_coeff2[t]] * vrai_coeff[t]

            t = t + 1

        t = 0

        the_final_result = 0

        while t < len(pre_final_l):
            the_final_result = the_final_result + pre_final_l[t]

            t = t + 1

        globalised_M = M_element_chem + the_final_result

        if coeff_stoechio != 0:
            globalised_M = globalised_M * coeff_stoechio

        if T < len(reactifs_f):

            r_equa.insert(T, globalised_M)

        else:

            r_equa2.insert(T, globalised_M)

        coeff_du_requa.insert(T, coeff_stoechio)

        T = T + 1

        print(r_equa)
